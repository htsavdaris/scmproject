from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd


def calculateSafetyStock():
    print('get vacc')


def getvaccCentersSypply():
    print('get vacc')


def getVaccCentersAppointments():
    #wb = Workbook()
    #wb = load_workbook(filename='E:\MYDOCS\05_Personal_Onedrive\OneDrive\98_Personal\22_COVID_Vacc\04_SCM_App\Master_Plan_v2.xlsm', read_only=True)
    #ws = wb['VaccCenters']
    #sheet_ranges = wb['range names']
    #print(sheet_ranges['D18'].value)
    #pd.read_excel('Master_Plan_v3.xlsx', sheet_name = 'VaccCenters', skiprows = 3,  nrows= 30, usecols = 'H:AF')
    pd.read_excel('Master_Plan_v3.xlsx', index_col=0)
    compression_opts = dict(method='zip',archive_name='out.csv')  
    pd.to_csv('out.zip', index=False, compression=compression_opts)